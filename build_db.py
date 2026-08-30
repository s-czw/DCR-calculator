#!/usr/bin/env python3
"""Build itc_rates.db from the ITC Trip & Parking Calculation Sheet.

    python3 build_db.py [path/to/ITC_Trip & Parking Calculation Sheet_V1.xls]

Reads the Rates-Weekdays and Rates-Weekend tabs into one `itc_rate` table keyed
by (period, code), and creates an `itc_combined` view carrying the weighted
week rate. Re-run after any reissue, then run embed_db.py to refresh the copy
embedded in dcr-calculator.html.

Legacy .xls needs xlrd:  pip install xlrd
.xlsx needs openpyxl.
"""
import os
import re
import sqlite3
import sys
from datetime import date

HERE = os.path.dirname(os.path.abspath(__file__))
DB = os.path.join(HERE, "itc_rates.db")
# The source workbooks are client data and are deliberately not in this repository.
# Drop them in ./sources (gitignored), point DCR_SOURCE_DIR at wherever they live,
# or pass the two paths as arguments.
SRC = os.environ.get("DCR_SOURCE_DIR") or os.path.join(HERE, "sources")
DEFAULT_XLS = os.path.join(SRC, "ITC_Trip & Parking Calculation Sheet_V1.xls")
DEFAULT_DESIG = os.path.join(SRC, "Designation_Index.xlsx")
DEFAULT_DED = os.path.join(SRC, "DED \u0642\u0627\u0626\u0645\u0629 \u0627\u0644\u0627\u0646\u0634\u0637\u0629 (2026 \u064a\u0648\u0644\u064a\u0648).xlsx")
DEFAULT_PLOTS = os.path.join(SRC, "DCR_plots.xlsx")
# One workbook per plot, named Plot_<sector/plot id>.xls with spaces written as
# underscores. Client data, so it lives outside the repository like the rest.
DEFAULT_UAD = os.environ.get("DCR_UAD_DIR") or os.path.join(HERE, "data",
                                                            "20260824_Full_20_Plots")
# Row 5 onwards; rows 1-4 are a disclaimer and a blank.
UAD_HEADER_ROW = 4
DEFAULT_UAD_MAP = os.path.join(SRC, "uad.xlsx")

# The mapping sheet names ITC land uses in its own words. These are the classes it
# means. Six differ only by plural or punctuation; two needed deciding:
#   Nurseries/Child Care -> 511, the only nursery class in the matrix.
#   Sport Centre         -> 632 Sports Club rather than 634 Special Sport Centre.
#     The rows carrying it are a fitness centre, indoor recreation and a jiu jitsu
#     club -- clubs rather than a special facility. THIS IS A JUDGEMENT, and the
#     entry here most worth a second opinion.
ITC_EQUIVALENT = {
    "on-street shopping":                 "115",
    "local shopping centre":              "112",
    "supermarkets":                       "114",
    "quality/high turnover restaurants":  "122",
    "fast food restaurants":              "123",
    "nurseries/child care":               "511",
    "private clinics":                    "822",
    "sport centre":                       "632",
}
UAD_COLS = ["ACTIVITYCODE", "MAIN_ACTIVITY_NAME", "ACTIVITY_VARIANTS_EN",
            "ACTIVITY_VARIANTS_Ar", "LICENSETYPEEN", "UAD_LUC_Code", "UAD_Category",
            "UAD_Main_Transportation_Mode", "Activity_Found_Walking_700m",
            "Activity_Found_Driving_10min", "UAD_Activity_Inclusion_Exclusion",
            "Plot_Proposed_UAD", "Rational", "Identifier", "IdentifeirLogic"]

# A standard week: five weekdays, two weekend days.
WEEK = {"weekday": 5.0, "weekend": 2.0}

# Sheet name -> period key. Add rows here if a reissue renames the tabs.
SHEETS = {"Rates-Weekdays": "weekday", "Rates-Weekend": "weekend",
          "Rates-Weekends": "weekend", "weekday": "weekday", "weekend": "weekend"}

# Column indices in the rate tabs (0-based), from the header on row 4.
COL = dict(cls=0, name=1, code=2, sites=3, am=5, am_in=6, noon=7, noon_in=8,
           pm=9, pm_in=10, phg=11, phg_in=12, emp=13, vis=14, trk=15,
           unit=16, driver=17, conv=18)
HEADER_ROWS = 4

DRIVER_KIND = {
    "GFA in Square Meter": "gfa",
    "GLA in Square Meter": "gla",
    "Total site area in Square Meter": "site",
}

# Navigational groupings only -- the workbook has no group names. Derived from
# the leading digit of the class code to keep the picker usable.
GROUPS = {
    "1": "Retail, Food & Showrooms",
    "2": "Government, Office & Services",
    "3": "Residential",
    "4": "Hotels & Serviced Apartments",
    "5": "Education, Culture & Worship",
    "6": "Leisure & Community",
    "7": "Industry & Logistics",
    "8": "Health",
    "9": "Transport",
}

SCHEMA = """
DROP VIEW  IF EXISTS itc_combined;
DROP TABLE IF EXISTS itc_rate;
DROP TABLE IF EXISTS weighting;
DROP TABLE IF EXISTS ded_activity;
DROP TABLE IF EXISTS land_use;
DROP TABLE IF EXISTS ded_code;
DROP TABLE IF EXISTS coefficient;
DROP TABLE IF EXISTS meta;

CREATE TABLE itc_rate (
    id             INTEGER PRIMARY KEY,
    period         TEXT NOT NULL CHECK (period IN ('weekday','weekend')),
    class_code     TEXT NOT NULL,
    class_name     TEXT NOT NULL,
    class_group    TEXT NOT NULL,
    code           TEXT NOT NULL,
    variant        TEXT,
    sites          REAL,              -- surveyed sites behind the rate
    trip_am        REAL, trip_am_in   REAL,
    trip_noon      REAL, trip_noon_in REAL,
    trip_pm        REAL, trip_pm_in   REAL,
    trip_phg       REAL, trip_phg_in  REAL,
    rate_employee  REAL NOT NULL,
    rate_visitor   REAL NOT NULL,
    rate_truck     REAL NOT NULL,
    rate_total     REAL NOT NULL,
    unit_label     TEXT NOT NULL,
    driver_label   TEXT NOT NULL,
    driver_kind    TEXT NOT NULL CHECK (driver_kind IN ('gfa','gla','site','count')),
    conversion     REAL NOT NULL,
    UNIQUE (period, code)
);
CREATE INDEX idx_itc_class  ON itc_rate (class_code);
CREATE INDEX idx_itc_group  ON itc_rate (class_group);
CREATE INDEX idx_itc_period ON itc_rate (period);

-- Days per week behind each period, and the share that implies.
CREATE TABLE weighting (
    period TEXT PRIMARY KEY,
    days   REAL NOT NULL,
    share  REAL NOT NULL
);

-- Base-district designations: the land-use selection, and the source of both
-- Maximum FAR and Maximum Plot Coverage. Coverage is a FRACTION (0.6 = 60%),
-- exactly as the Code publishes it. far/coverage are NULL where the Code
-- publishes no number -- the reason is in remarks.
CREATE TABLE land_use (
    id          INTEGER PRIMARY KEY,
    category    TEXT NOT NULL,
    designation TEXT NOT NULL,
    name        TEXT NOT NULL,
    far         REAL,
    coverage    REAL,
    remarks     TEXT,
    source      TEXT NOT NULL DEFAULT 'code'
);
CREATE INDEX idx_lu_desig ON land_use (designation);

-- DED economic activities, one row per distinct active ACTIVITY_ID.
-- itc_class is a HEURISTIC crosswalk authored here, not a client-supplied
-- mapping: see ITC_RULES in this script. map_confidence says how much to trust
-- it, and the calculator lets any activity be reassigned by hand.
CREATE TABLE ded_activity (
    activity_id    TEXT PRIMARY KEY,
    name           TEXT NOT NULL,
    section        TEXT,
    division       TEXT,
    isic_class     TEXT,
    nature         TEXT,
    itc_class      TEXT,
    map_confidence TEXT CHECK (map_confidence IN ('high','medium','low')),
    map_reason     TEXT
);
CREATE INDEX idx_act_div  ON ded_activity (division);
CREATE INDEX idx_act_itc  ON ded_activity (itc_class);

-- The plot register: the plots a user can pick instead of typing an area.
-- sector_plot_id is the address the saved-configuration JSON is keyed by, so it
-- is UNIQUE and NOT NULL -- two plots sharing one would silently overwrite each
-- other on save.
CREATE TABLE plot (
    id            INTEGER PRIMARY KEY,
    plot_number   TEXT,
    sector_plot_id TEXT NOT NULL UNIQUE,
    district      TEXT,
    primary_use   TEXT,
    devcode       TEXT,
    devcode_cat   TEXT,
    area          REAL
);
CREATE INDEX idx_plot_sector ON plot (sector_plot_id);

-- The UAD code/category the user picks from, and the ITC class whose rate applies.
CREATE TABLE uad_map (
    uad_code       TEXT PRIMARY KEY,
    uad_category   TEXT NOT NULL,
    itc_equivalent TEXT NOT NULL,
    itc_class      TEXT NOT NULL,
    note           TEXT
);

-- Per-plot UAD activities, one row per activity that passed BOTH filters:
--   Plot_Proposed_UAD = Yes   AND   UAD_Activity_Inclusion_Exclusion = Include
-- The excluded rows are not loaded; the workbooks stay outside the repository as
-- the record of what was filtered out.
CREATE TABLE plot_uad (
    id              INTEGER PRIMARY KEY,
    sector_plot_id  TEXT NOT NULL,
    activity_code   TEXT,
    activity_name   TEXT,
    variant_en      TEXT,
    variant_ar      TEXT,
    licence_type    TEXT,
    luc_code        TEXT,
    category        TEXT,
    transport_mode  TEXT,
    found_walk_700m TEXT,
    found_drive_10m TEXT,
    inclusion       TEXT,
    proposed        TEXT,
    rational        TEXT,
    identifier      TEXT,
    identifier_logic TEXT
);
CREATE INDEX idx_uad_plot ON plot_uad (sector_plot_id);

CREATE TABLE coefficient (key TEXT PRIMARY KEY, value REAL NOT NULL, note TEXT);
CREATE TABLE meta (key TEXT PRIMARY KEY, value TEXT NOT NULL);
"""

# The weighted week rate, as a view so it is queryable rather than only in the UI.
VIEW = """
CREATE VIEW itc_combined AS
SELECT
    wd.class_code, wd.class_name, wd.class_group, wd.code, wd.variant,
    wd.unit_label, wd.driver_label, wd.driver_kind, wd.conversion,
    wd.rate_total AS weekday_total,
    we.rate_total AS weekend_total,
    ROUND(wd.rate_employee * ws.share + we.rate_employee * es.share, 6) AS rate_employee,
    ROUND(wd.rate_visitor  * ws.share + we.rate_visitor  * es.share, 6) AS rate_visitor,
    ROUND(wd.rate_truck    * ws.share + we.rate_truck    * es.share, 6) AS rate_truck,
    ROUND(wd.rate_total    * ws.share + we.rate_total    * es.share, 6) AS rate_total
FROM itc_rate wd
JOIN itc_rate we ON we.code = wd.code AND we.period = 'weekend'
JOIN weighting ws ON ws.period = 'weekday'
JOIN weighting es ON es.period = 'weekend'
WHERE wd.period = 'weekday';
"""


def load_sheets(path):
    """Return {period: [row dicts]} from .xls (xlrd) or .xlsx (openpyxl)."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        try:
            import xlrd
        except ImportError:
            sys.exit("Reading .xls needs xlrd:  pip install xlrd")
        wb = xlrd.open_workbook(path)
        names = wb.sheet_names()
        get = lambda nm: [[wb.sheet_by_name(nm).cell_value(r, c)
                           for c in range(wb.sheet_by_name(nm).ncols)]
                          for r in range(wb.sheet_by_name(nm).nrows)]
    else:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        names = wb.sheetnames
        get = lambda nm: [list(r) for r in wb[nm].iter_rows(values_only=True)]

    found = {}
    for nm in names:
        period = SHEETS.get(nm.strip())
        if period:
            found[period] = get(nm)
    missing = set(WEEK) - set(found)
    if missing:
        sys.exit(f"no sheet for period(s) {sorted(missing)}; sheets present: {names}")
    return found


def load_designations(path):
    """Read the Designation Index into land_use rows.

    Prefers the 'Flat Data' sheet, which carries the category on every row; the
    'Designations' sheet uses category banner rows instead.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    rows = []
    if "Flat Data" in wb.sheetnames:
        grid = [list(r) for r in wb["Flat Data"].iter_rows(values_only=True)]
        for r in grid[1:]:
            if not r or not r[1]:
                continue
            rows.append(dict(category=str(r[0] or "").strip(),
                             designation=str(r[1]).strip(),
                             name=str(r[2] or "").strip(),
                             far=numeric(r[3]), coverage=numeric(r[4]),
                             remarks=(str(r[5]).strip() if len(r) > 5 and r[5] else None)))
    else:
        grid = [list(r) for r in wb["Designations"].iter_rows(values_only=True)]
        cat = ""
        for r in grid[1:]:
            if not r or not r[0]:
                continue
            # a banner row carries only the category name
            if all(x in (None, "") for x in r[1:4]):
                cat = str(r[0]).strip()
                continue
            rows.append(dict(category=cat, designation=str(r[0]).strip(),
                             name=str(r[1] or "").strip(),
                             far=numeric(r[2]), coverage=numeric(r[3]),
                             remarks=(str(r[4]).strip() if len(r) > 4 and r[4] else None)))
    if not rows:
        sys.exit(f"no designation rows parsed from {path}")
    bad = [r["designation"] for r in rows
           if r["coverage"] is not None and not 0 < r["coverage"] <= 1]
    if bad:
        sys.exit(f"coverage outside (0,1] -- is it a percentage not a fraction? {bad}")
    return rows


# Heuristic DED -> ITC crosswalk. Ordered; first match wins. Keyed on the
# numeric ISIC levels because the workbook's LEVEL_0 letters are re-lettered and
# do not follow ISIC sections. Confidence is deliberately conservative: "high"
# means the ISIC class and the ITC category describe the same thing, "low" means
# the ITC category is the closest available analogue and should be reviewed.
#   kind: name3 = keyword within a given 4-digit class
#         l3    = exact 4-digit ISIC class
#         div   = exact 2-digit division
#         range = inclusive division range
ITC_RULES = [
    ("name3", ("5610", r"fast food|cafeteria|caf\u00e9|cafe\b|coffee|snack|juice|shawarma|kiosk"),
     "123", "high",   "food service, quick-service format"),
    ("l3",    "5610", "122", "medium", "restaurant, assumed quality/high-turnover"),
    ("div",   "56",   "122", "medium", "food and beverage service"),

    ("name3", ("4772", r"pharmac"), "831", "high",   "pharmaceutical retail"),
    ("l3",    "4711", "114", "high",   "non-specialised food retail = supermarket"),
    ("l3",    "4719", "113", "high",   "other non-specialised retail = superstore"),
    ("l3",    "4741", "142", "high",   "computer and telecom retail"),
    ("l3",    "4742", "142", "high",   "audio/video equipment retail"),
    ("l3",    "4759", "141", "high",   "furniture and household equipment retail"),
    ("div",   "47",   "112", "medium", "specialised retail, assumed local shopping centre"),
    ("div",   "45",   "144", "medium", "motor vehicle trade = vehicle showroom"),
    ("div",   "46",   "112", "low",    "wholesale trade; shopfront assumed, review if warehousing"),

    ("div",   "64",   "235", "high",   "financial service = banking service"),
    ("div",   "65",   "235", "high",   "insurance, treated as banking service"),
    ("div",   "66",   "235", "high",   "auxiliary financial service"),

    ("l3",    "8510", "516", "high",   "kindergarten"),
    ("div",   "85",   "515", "medium", "education, assumed private school"),
    ("l3",    "8610", "812", "high",   "hospital activities"),
    ("div",   "86",   "822", "medium", "human health, assumed private clinic"),
    ("div",   "87",   "822", "low",    "residential care, closest analogue is clinic"),
    ("div",   "88",   "822", "low",    "social work, closest analogue is clinic"),

    ("div",   "55",   "411", "medium", "accommodation"),
    ("div",   "84",   "211", "medium", "public administration"),
    ("div",   "93",   "632", "medium", "sports and recreation"),
    ("div",   "90",   "631", "low",    "creative and arts, closest analogue is social club"),
    ("div",   "91",   "532", "low",    "libraries and museums"),
    ("div",   "92",   "631", "low",    "gambling, closest analogue is social club"),

    ("range", (5, 9),   "731", "medium", "mining and quarrying = heavy industry"),
    ("range", (10, 33), "711", "medium", "manufacturing = production oriented industry"),
    ("range", (35, 39), "711", "low",    "utilities and waste, closest analogue is industry"),
    ("range", (41, 43), "222", "low",    "construction, assumed contractor office"),
    ("range", (49, 53), "713", "medium", "transport and storage = warehousing"),
    ("range", (58, 63), "222", "medium", "information and communication = office"),
    ("range", (68, 75), "222", "medium", "real estate and professional services = office"),
    ("range", (77, 82), "222", "medium", "administrative and support services = office"),
    ("range", (94, 96), "112", "low",    "membership, repair and personal services; shopfront assumed"),
    ("range", (1, 3),   "711", "low",    "agriculture, closest analogue is industry"),
]
FALLBACK = ("112", "low", "no rule matched; defaulted to local shopping centre retail")


def map_activity(division, isic_class, name):
    """DED activity -> (itc_class, confidence, reason)."""
    low = (name or "").lower()
    for kind, val, cls, conf, why in ITC_RULES:
        if kind == "name3":
            l3, pattern = val
            if isic_class == l3 and re.search(pattern, low):
                return cls, conf, why
        elif kind == "l3":
            if isic_class == val:
                return cls, conf, why
        elif kind == "div":
            if division == val:
                return cls, conf, why
        elif kind == "range":
            lo, hi = val
            if division.isdigit() and lo <= int(division) <= hi:
                return cls, conf, why
    return FALLBACK


def load_activities(path):
    """Distinct active DED activities. ACTIVITY_ID repeats across sub-natures,
    so keep the first occurrence of each."""
    import openpyxl
    ws = openpyxl.load_workbook(path, data_only=True, read_only=True)["DED"]
    it = ws.iter_rows(values_only=True)
    hdr = next(it)
    C = {h: i for i, h in enumerate(hdr)}
    need = ("ACTIVITY_ID", "ACTIVITY_NAME_EN", "ACTIVITY_STATUS_EN",
            "LEVEL_0", "LEVEL_1", "LEVEL_3", "NATURE_EN")
    for k in need:
        if k not in C:
            sys.exit(f"DED sheet is missing column {k}; columns present: {list(C)[:8]}...")
    out = {}
    for r in it:
        aid = norm(r[C["ACTIVITY_ID"]])
        if not aid or aid in out:
            continue
        if norm(r[C["ACTIVITY_STATUS_EN"]]) != "Active":
            continue
        name = norm(r[C["ACTIVITY_NAME_EN"]])
        div = norm(r[C["LEVEL_1"]])
        l3 = norm(r[C["LEVEL_3"]])
        cls, conf, why = map_activity(div, l3, name)
        out[aid] = dict(activity_id=aid, name=name, section=norm(r[C["LEVEL_0"]]),
                        division=div, isic_class=l3, nature=norm(r[C["NATURE_EN"]]),
                        itc_class=cls, map_confidence=conf, map_reason=why)
    return list(out.values())


def norm(v):
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def numeric(v):
    return float(v) if isinstance(v, (int, float)) else None


def parse(grid):
    """Rate rows from one tab. Class and Class Name are merged across a class's
    variant rows, so both carry forward until the next non-empty cell."""
    out, cls, nm = [], "", ""
    for row in grid[HEADER_ROWS:]:
        cell = lambda k: row[COL[k]] if COL[k] < len(row) else None
        if norm(cell("cls")):
            cls = norm(cell("cls"))
        if norm(cell("name")):
            nm = norm(cell("name"))
        code = norm(cell("code"))
        emp = numeric(cell("emp"))
        if not code or emp is None:
            continue                      # spacer or section row
        vis = numeric(cell("vis")) or 0.0
        trk = numeric(cell("trk")) or 0.0
        driver = norm(cell("driver"))
        variant = code[len(cls) + 1:] if code.startswith(cls + "-") else None
        out.append(dict(
            class_code=cls, class_name=nm, class_group=GROUPS.get(cls[:1], "Other"),
            code=code, variant=variant, sites=numeric(cell("sites")),
            trip_am=numeric(cell("am")), trip_am_in=numeric(cell("am_in")),
            trip_noon=numeric(cell("noon")), trip_noon_in=numeric(cell("noon_in")),
            trip_pm=numeric(cell("pm")), trip_pm_in=numeric(cell("pm_in")),
            trip_phg=numeric(cell("phg")), trip_phg_in=numeric(cell("phg_in")),
            rate_employee=emp, rate_visitor=vis, rate_truck=trk,
            rate_total=round(emp + vis + trk, 6),
            unit_label=norm(cell("unit")), driver_label=driver,
            driver_kind=DRIVER_KIND.get(driver, "count"),
            conversion=numeric(cell("conv")) or 1.0,
        ))
    return out


def load_uad_map(path):
    """UAD code -> UAD category -> the ITC class its rate comes from.

    Row 1 groups the columns, row 2 names them, data starts at row 3. The sheet
    holds two "UADs Categories" columns -- one according to Inspection, one
    according to LUC. The LUC one is taken, because the plot workbooks report
    UAD_LUC_Code and UAD_Category on that same basis.
    """
    from openpyxl import load_workbook
    ws = load_workbook(path, data_only=True).worksheets[0]
    rows = [[("" if v is None else str(v).strip()) for v in r]
            for r in ws.iter_rows(min_row=3, values_only=True)]
    out, seen = [], {}
    for r in rows:
        if len(r) < 4 or not any(r):
            continue
        code = r[1].split(".")[0].strip()
        cat, equiv = r[2].strip(), r[3].strip()
        if not code or code == "0" or not equiv:
            continue          # an unnumbered row cannot be matched to a plot row
        itc = ITC_EQUIVALENT.get(equiv.lower())
        if itc is None:
            sys.exit(f"{os.path.basename(path)}: ITC land use {equiv!r} is not in "
                     f"ITC_EQUIVALENT; add it there with the class it means")
        if code in seen:
            if seen[code] != (cat, itc):
                sys.exit(f"{os.path.basename(path)}: UAD code {code} maps two ways: "
                         f"{seen[code]} and {(cat, itc)}")
            continue
        seen[code] = (cat, itc)
        out.append(dict(uad_code=code, uad_category=cat, itc_equivalent=equiv,
                        itc_class=itc, note=r[4].strip() if len(r) > 4 else None))
    return out


def load_uad(folder, known_ids):
    """Per-plot UAD activities, keeping only rows that pass both filters."""
    import glob
    if not folder or not os.path.isdir(folder):
        return [], []
    import xlrd
    # The filename writes spaces as underscores; the register keeps the spaces, and
    # keeps a trailing underscore where the plot id has one. Normalise to compare.
    lookup = {r.replace(" ", "_"): r for r in known_ids}
    out, unmatched = [], []
    for f in sorted(glob.glob(os.path.join(folder, "Plot_*.xls"))):
        stem = os.path.basename(f)[len("Plot_"):-len(".xls")]
        plot_id = lookup.get(stem)
        if plot_id is None:
            unmatched.append(stem)
            plot_id = stem          # keep the data; it simply cannot be picked yet
        sh = xlrd.open_workbook(f).sheet_by_index(0)
        hdr = [norm(sh.cell_value(UAD_HEADER_ROW, c)) or "" for c in range(sh.ncols)]
        ix = {name: hdr.index(name) for name in UAD_COLS if name in hdr}
        for want in ("Plot_Proposed_UAD", "UAD_Activity_Inclusion_Exclusion"):
            if want not in ix:
                sys.exit(f"{os.path.basename(f)}: no {want} column; headers were {hdr}")
        for r in range(UAD_HEADER_ROW + 1, sh.nrows):
            cell = lambda n: norm(sh.cell_value(r, ix[n])) if n in ix else None
            if (cell("Plot_Proposed_UAD") or "").lower() != "yes":
                continue
            if (cell("UAD_Activity_Inclusion_Exclusion") or "").lower() != "include":
                continue
            out.append(dict(
                sector_plot_id=plot_id, activity_code=cell("ACTIVITYCODE"),
                activity_name=cell("MAIN_ACTIVITY_NAME"), variant_en=cell("ACTIVITY_VARIANTS_EN"),
                variant_ar=cell("ACTIVITY_VARIANTS_Ar"), licence_type=cell("LICENSETYPEEN"),
                luc_code=cell("UAD_LUC_Code"), category=cell("UAD_Category"),
                transport_mode=cell("UAD_Main_Transportation_Mode"),
                found_walk_700m=cell("Activity_Found_Walking_700m"),
                found_drive_10m=cell("Activity_Found_Driving_10min"),
                inclusion=cell("UAD_Activity_Inclusion_Exclusion"),
                proposed=cell("Plot_Proposed_UAD"), rational=cell("Rational"),
                identifier=cell("Identifier"), identifier_logic=cell("IdentifeirLogic")))
    return out, unmatched


def load_plots(path):
    """The plot register: one row per plot, keyed by its sector/plot address."""
    from openpyxl import load_workbook
    ws = load_workbook(path, data_only=True).worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    # Match columns by header text rather than position: the register is
    # hand-maintained and a reissue that adds a column should not silently shift
    # the area into the DevCode field.
    head = [norm(h).lower() if h is not None else "" for h in rows[0]]
    def col(*names):
        for want in names:
            for i, h in enumerate(head):
                if h == want:
                    return i
        for want in names:
            for i, h in enumerate(head):
                if want in h:
                    return i
        return None
    ix = {
        "plot_number":    col("plot number"),
        "sector_plot_id": col("sector / plot id", "sector/plot id", "sector"),
        "district":       col("district"),
        "primary_use":    col("primary use"),
        "devcode":        col("devcode"),
        "devcode_cat":    col("devcode category"),
        "area":           col("plot area (m2)", "plot area"),
    }
    for key in ("sector_plot_id", "area"):
        if ix[key] is None:
            sys.exit(f"{os.path.basename(path)}: no column found for {key}; "
                     f"headers were {head}")
    # "DevCode" is a prefix of "DevCode category", so a substring match could pick
    # the wrong one. Reject that rather than mis-file every row.
    if ix["devcode"] is not None and ix["devcode"] == ix["devcode_cat"]:
        ix["devcode_cat"] = None

    out, seen = [], {}
    for r in rows[1:]:
        def cell(key):
            i = ix[key]
            return norm(r[i]) if i is not None and i < len(r) else None
        sector = cell("sector_plot_id")
        if not sector:
            continue
        area = numeric(r[ix["area"]]) if ix["area"] < len(r) else None
        if sector in seen:
            sys.exit(f"{os.path.basename(path)}: sector/plot id {sector!r} appears "
                     f"twice (rows {seen[sector]} and {len(out) + 2}). It is the key "
                     f"saved configurations are stored under, so it must be unique.")
        seen[sector] = len(out) + 2
        out.append(dict(plot_number=cell("plot_number"), sector_plot_id=sector,
                        district=cell("district"), primary_use=cell("primary_use"),
                        devcode=cell("devcode"), devcode_cat=cell("devcode_cat"),
                        area=area))
    return out


UAD_MAP_DDL = """
CREATE TABLE uad_map (
    uad_code       TEXT PRIMARY KEY,
    uad_category   TEXT NOT NULL,
    itc_equivalent TEXT NOT NULL,
    itc_class      TEXT NOT NULL,
    note           TEXT
);
"""

UAD_MAP_INSERT = ("INSERT INTO uad_map (uad_code, uad_category, itc_equivalent, itc_class, "
                  "note) VALUES (:uad_code,:uad_category,:itc_equivalent,:itc_class,:note)")

UAD_DDL = """
CREATE TABLE plot_uad (
    id              INTEGER PRIMARY KEY,
    sector_plot_id  TEXT NOT NULL,
    activity_code   TEXT,
    activity_name   TEXT,
    variant_en      TEXT,
    variant_ar      TEXT,
    licence_type    TEXT,
    luc_code        TEXT,
    category        TEXT,
    transport_mode  TEXT,
    found_walk_700m TEXT,
    found_drive_10m TEXT,
    inclusion       TEXT,
    proposed        TEXT,
    rational        TEXT,
    identifier      TEXT,
    identifier_logic TEXT
);
CREATE INDEX idx_uad_plot ON plot_uad (sector_plot_id);
"""

UAD_INSERT = (
    "INSERT INTO plot_uad (sector_plot_id, activity_code, activity_name, variant_en, "
    "variant_ar, licence_type, luc_code, category, transport_mode, found_walk_700m, "
    "found_drive_10m, inclusion, proposed, rational, identifier, identifier_logic) VALUES "
    "(:sector_plot_id,:activity_code,:activity_name,:variant_en,:variant_ar,:licence_type,"
    ":luc_code,:category,:transport_mode,:found_walk_700m,:found_drive_10m,:inclusion,"
    ":proposed,:rational,:identifier,:identifier_logic)")


def uad_only(folder, map_path=None):
    """Refresh just plot_uad on the existing database.

    The UAD workbooks are reissued on their own schedule, and the rate matrix and
    DED list they would otherwise be rebuilt alongside are client files that are
    not kept here. Rebuilding everything to add one table would mean needing all
    of them present; this needs only the workbooks that changed.
    """
    if not os.path.exists(DB):
        sys.exit(f"{DB} not found - a full build has to run first")
    con = sqlite3.connect(DB)
    con.row_factory = sqlite3.Row
    known = [r["sector_plot_id"] for r in con.execute("SELECT sector_plot_id FROM plot")]
    rows, unmatched = load_uad(folder, known)
    if not rows:
        sys.exit(f"no UAD rows found in {folder}")

    map_path = map_path or DEFAULT_UAD_MAP
    mapping = load_uad_map(map_path) if os.path.exists(map_path) else []
    if mapping:
        # An ITC class the matrix does not hold would give a silent zero rate.
        have = {r[0] for r in con.execute("SELECT DISTINCT class_code FROM itc_rate")}
        orphan = sorted({m["itc_class"] for m in mapping if m["itc_class"] not in have})
        if orphan:
            sys.exit(f"uad_map points at ITC classes absent from the matrix: {orphan}")
        con.executescript("DROP TABLE IF EXISTS uad_map;")
        con.executescript(UAD_MAP_DDL)
        con.executemany(UAD_MAP_INSERT, mapping)
        con.execute("INSERT OR REPLACE INTO meta (key, value) VALUES (?,?)",
                    ("uad_map_rows", str(len(mapping))))
        print(f"  {len(mapping)} UAD codes mapped to ITC classes")
        # Which of the plots' codes the mapping does not reach.
        codes = {m["uad_code"] for m in mapping}
        used = {}
        for r in rows:
            c = (r["luc_code"] or "").split(".")[0].strip()
            used[c] = used.get(c, 0) + 1
        # A workbook row can arrive with no LUC code at all. Where its category
        # names a mapping row exactly, that is the code it meant -- the v2 sheet
        # gave Jiu Jitsu club the code 7240 while the plot workbooks still carry a
        # blank for it. Filled in here so the schedule resolves rather than asking
        # for an ITC class by hand.
        by_cat = {m["uad_category"].strip().lower(): m["uad_code"] for m in mapping}
        healed = 0
        for r in rows:
            if (r["luc_code"] or "").strip():
                continue
            hit = by_cat.get((r["category"] or "").strip().lower())
            if hit:
                r["luc_code"] = hit
                healed += 1
        if healed:
            print(f"  {healed} rows had no LUC code and were matched on their category")
            con.executescript("DELETE FROM plot_uad;")
            con.executemany(UAD_INSERT, rows)
            used = {}
            for r in rows:
                c = (r["luc_code"] or "").split(".")[0].strip()
                used[c] = used.get(c, 0) + 1
        gaps = {c: n for c, n in used.items() if c not in codes}
        if gaps:
            total = sum(gaps.values())
            print(f"  {total} activity rows have no mapped UAD code "
                  f"({len(gaps)} distinct): {sorted(gaps)[:6]}")
            print(f"    they are kept and shown, but need an ITC class chosen by hand")
    con.executescript("DROP TABLE IF EXISTS plot_uad;")
    con.executescript(UAD_DDL)
    con.executemany(UAD_INSERT, rows)
    con.execute("INSERT OR REPLACE INTO meta (key, value) VALUES (?,?)",
                ("uad_dir", os.path.basename(os.path.normpath(folder))))
    con.execute("INSERT OR REPLACE INTO meta (key, value) VALUES (?,?)",
                ("uad_rows", str(len(rows))))
    n_plots = len({r["sector_plot_id"] for r in rows})
    con.execute("INSERT OR REPLACE INTO meta (key, value) VALUES (?,?)",
                ("uad_plots", str(n_plots)))
    con.commit()
    print(f"  {len(rows):,} UAD activities kept across {n_plots} plots")
    print(f"  filters: Plot_Proposed_UAD = Yes and UAD_Activity_Inclusion_Exclusion = Include")
    if unmatched:
        print(f"  no register entry for {unmatched} - loaded under the filename id, but the "
              f"plot cannot be selected until it is added to DCR_plots.xlsx")
    return len(rows)


def main():
    if "--uad-only" in sys.argv:
        rest = [a for a in sys.argv[1:] if a != "--uad-only"]
        uad_only(rest[0] if rest else DEFAULT_UAD,
                 rest[1] if len(rest) > 1 else None)
        return

    xls = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLS
    if not os.path.exists(xls):
        sys.exit(f"not found: {xls}")
    sheets = load_sheets(xls)

    con = sqlite3.connect(DB)
    con.executescript(SCHEMA)
    total_days = sum(WEEK.values())
    con.executemany("INSERT INTO weighting (period, days, share) VALUES (?,?,?)",
                    [(p, d, d / total_days) for p, d in WEEK.items()])

    counts = {}
    fields = ("period,class_code,class_name,class_group,code,variant,sites,"
              "trip_am,trip_am_in,trip_noon,trip_noon_in,trip_pm,trip_pm_in,"
              "trip_phg,trip_phg_in,rate_employee,rate_visitor,rate_truck,"
              "rate_total,unit_label,driver_label,driver_kind,conversion")
    keys = fields.split(",")[1:]
    for period, grid in sheets.items():
        rows = parse(grid)
        counts[period] = len(rows)
        con.executemany(
            f"INSERT INTO itc_rate ({fields}) VALUES ({','.join(['?'] * (len(keys) + 1))})",
            [tuple([period] + [r[k] for k in keys]) for r in rows])
    con.executescript(VIEW)

    # A code present in one tab but not the other would silently drop out of the
    # combined view, so fail loudly instead.
    orphans = con.execute("""
        SELECT period, code FROM itc_rate
        WHERE code NOT IN (SELECT code FROM itc_rate WHERE period='weekday')
           OR code NOT IN (SELECT code FROM itc_rate WHERE period='weekend')
        ORDER BY code""").fetchall()
    if orphans:
        sys.exit(f"codes missing from one tab, combined view would drop them: {orphans}")

    desig = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_DESIG
    lu = load_designations(desig) if os.path.exists(desig) else []

    ded = sys.argv[3] if len(sys.argv) > 3 else DEFAULT_DED
    plots_src = sys.argv[4] if len(sys.argv) > 4 else DEFAULT_PLOTS
    plots = load_plots(plots_src) if os.path.exists(plots_src) else []
    uad_dir = sys.argv[5] if len(sys.argv) > 5 else DEFAULT_UAD
    uad, uad_unmatched = load_uad(uad_dir, [p["sector_plot_id"] for p in plots])
    acts = load_activities(ded) if os.path.exists(ded) else []
    con.executemany(
        "INSERT INTO ded_activity (activity_id, name, section, division, isic_class, "
        "nature, itc_class, map_confidence, map_reason) VALUES "
        "(:activity_id,:name,:section,:division,:isic_class,:nature,:itc_class,"
        ":map_confidence,:map_reason)", acts)

    # An activity mapped to an ITC class that does not exist would silently give
    # no rate in the calculator, so fail the build instead.
    orphan_cls = con.execute("""
        SELECT DISTINCT a.itc_class FROM ded_activity a
        WHERE a.itc_class NOT IN (SELECT class_code FROM itc_rate)
        ORDER BY 1""").fetchall()
    if orphan_cls:
        sys.exit(f"ded_activity maps to ITC classes absent from the matrix: "
                 f"{[c[0] for c in orphan_cls]}")
    con.executemany(
        "INSERT INTO land_use (category, designation, name, far, coverage, remarks, source) "
        "VALUES (:category,:designation,:name,:far,:coverage,:remarks,'code')", lu)
    con.executemany(
        "INSERT INTO plot (plot_number, sector_plot_id, district, primary_use, "
        "devcode, devcode_cat, area) VALUES (:plot_number,:sector_plot_id,:district,"
        ":primary_use,:devcode,:devcode_cat,:area)", plots)
    con.executemany(
        "INSERT INTO plot_uad (sector_plot_id, activity_code, activity_name, variant_en, "
        "variant_ar, licence_type, luc_code, category, transport_mode, found_walk_700m, "
        "found_drive_10m, inclusion, proposed, rational, identifier, identifier_logic) VALUES "
        "(:sector_plot_id,:activity_code,:activity_name,:variant_en,:variant_ar,:licence_type,"
        ":luc_code,:category,:transport_mode,:found_walk_700m,:found_drive_10m,:inclusion,"
        ":proposed,:rational,:identifier,:identifier_logic)", uad)
    con.executemany("INSERT INTO coefficient (key, value, note) VALUES (?,?,?)", [
        ("gla_pct", 75, "GLA = Max GFA x this %. User-editable; not a Code citation."),
        ("coverage_pct_default", 60,
         "Fallback Max Plot Coverage % when the selected designation publishes none."),
    ])
    con.executemany("INSERT INTO meta (key, value) VALUES (?,?)", [
        ("source_file", os.path.basename(xls)),
        ("sheets", ", ".join(sorted(sheets))),
        ("rate_rows", str(sum(counts.values()))),
        ("codes", str(counts.get("weekday", 0))),
        ("week_split", " / ".join(f"{p} {int(d)}d" for p, d in WEEK.items())),
        ("designation_file", os.path.basename(desig) if lu else "(none)"),
        ("designations", str(len(lu))),
        ("activity_file", os.path.basename(ded) if acts else "(none)"),
        ("activities", str(len(acts))),
        ("plot_file", os.path.basename(plots_src) if plots else "(none)"),
        ("plots", str(len(plots))),
        ("uad_dir", os.path.basename(os.path.normpath(uad_dir)) if uad else "(none)"),
        ("uad_rows", str(len(uad))),
        ("uad_plots", str(len({r["sector_plot_id"] for r in uad}))),
        ("built", date.today().isoformat()),
    ])
    con.commit()

    if uad:
        n_plots = len({r["sector_plot_id"] for r in uad})
        print(f"  UAD: {len(uad):,} activities kept across {n_plots} plots "
              f"(Plot_Proposed_UAD = Yes and UAD_Activity_Inclusion_Exclusion = Include)")
    if uad_unmatched:
        print(f"  UAD workbooks with no plot in the register: {uad_unmatched} "
              f"- their activities are loaded under the filename id, but the plot "
              f"cannot be selected until it is added to the register")

    print(f"{DB}  <-  {os.path.basename(xls)}")
    for p in sorted(counts):
        print(f"  {p:8s} {counts[p]:4d} rows   share {WEEK[p]/total_days*100:6.2f}%")
    for kind, n in con.execute("SELECT driver_kind, COUNT(*) FROM itc_rate "
                              "WHERE period='weekday' GROUP BY 1 ORDER BY 2 DESC"):
        print(f"    {kind:6s} {n:4d}")
    n, = con.execute("SELECT COUNT(*) FROM itc_combined").fetchone()
    print(f"  itc_combined view: {n} codes")
    if lu:
        tot, wfar, wcov = con.execute(
            "SELECT COUNT(*), COUNT(far), COUNT(coverage) FROM land_use").fetchone()
        print(f"  land_use: {tot} designations  ({wfar} with FAR, {wcov} with coverage)")
        for cat, k in con.execute("SELECT category, COUNT(*) FROM land_use "
                                  "GROUP BY category ORDER BY MIN(id)"):
            print(f"    {cat:12s} {k:3d}")
    if acts:
        print(f"  ded_activity: {len(acts)} active activities")
        for conf, k in con.execute("SELECT map_confidence, COUNT(*) FROM ded_activity "
                                   "GROUP BY 1 ORDER BY CASE map_confidence "
                                   "WHEN 'high' THEN 1 WHEN 'medium' THEN 2 ELSE 3 END"):
            print(f"    {conf:8s} {k:5d}  ({k/len(acts)*100:4.1f}%)")
        print("    top ITC classes:")
        for cls, k, nm in con.execute(
                "SELECT itc_class, COUNT(*), "
                "  (SELECT class_name FROM itc_rate r WHERE r.class_code = a.itc_class LIMIT 1) "
                "FROM ded_activity a GROUP BY itc_class ORDER BY 2 DESC LIMIT 10"):
            print(f"      {cls:5s} {k:5d}  {str(nm)[:44]}")
    con.close()


if __name__ == "__main__":
    main()
