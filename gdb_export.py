#!/usr/bin/env python3
"""Run the DCR derivation over every plot in a file geodatabase, to Excel.

    python3 gdb_export.py DMT_Plot_Entrance.gdb [-o plots.xlsx]

Plot areas come from the UDM_Plot layer's PLOTCALCULATEDAREA. Floor area ratio
comes from the layer's own DevCode_FAR where it has one, falling back to the
designation schedule in itc_rates.db; every row records which source was used.
Maximum plot coverage is not in the geodatabase, so it always comes from the
schedule, or from the fallback default where the Code publishes none.

Needs GDAL through pyogrio, plus openpyxl:

    pip install pyogrio openpyxl

Parking is deliberately not calculated per plot. The number of spaces required
depends on which activities occupy the plot, which no geodatabase field carries.
What the sheet gives instead is the envelope: how many spaces the open ground
outside the footprint holds, how much one basement floor yields, and the
ceiling limitations L-3/L-4 put on demand.
"""
from __future__ import annotations

import argparse
import math
import os
import sqlite3
import sys
from datetime import datetime

HERE = os.path.dirname(os.path.abspath(__file__))
DB = os.path.join(HERE, "itc_rates.db")

# Limitations L-3 and L-4 of the Development Control Regulations cap parking
# demand at 2 spaces per 100 sqm of GFA. Above the cap: L-3 makes the excess the
# developer's responsibility on a plot of 1,600 sqm or more, and L-4 bars it
# outright below that. Kept in step with PARK_CAP/SMALL_PLOT in the calculator.
PARK_CAP = 2.0
SMALL_PLOT = 1600.0

# Fields read from the plot layer. Everything is optional except the area.
AREA_FIELD = "PLOTCALCULATEDAREA"
WANTED = [
    "PLOTNUMBER", "SectorNumber_PlotID", "GISID", "DISTRICTENG", "COMMUNITYENG",
    "PRIMARYUSEENGDESC", "SECUSEENG", AREA_FIELD,
    "DevCode", "DevCode_Category", "DevCode_Description", "DevCode_FAR", "DevCode_MaxGFA",
]

COLUMNS = [
    ("#",                          6,  None),
    ("Plot number",               14,  None),
    ("Sector / plot ID",          20,  None),
    ("District",                  20,  None),
    ("Primary use",               22,  None),
    ("DevCode",                   14,  None),
    ("DevCode category",          20,  None),
    ("Plot area (m2)",            14,  "#,##0.00"),
    ("FAR",                        8,  "0.000"),
    ("FAR source",                12,  None),
    ("Max GFA (m2)",              14,  "#,##0.00"),
    ("gdb Max GFA (m2)",          16,  "#,##0.00"),
    ("GFA check",                 12,  None),
    ("Max plot coverage",         17,  "0.0%"),
    ("Coverage source",           16,  None),
    ("Plot coverage (m2)",        17,  "#,##0.00"),
    ("GLA (m2)",                  13,  "#,##0.00"),
    ("Unit area (m2)",            13,  "#,##0"),
    ("Max activities",            13,  "#,##0"),
    ("Open ground (m2)",          17,  "#,##0.00"),
    ("Spaces on open ground",     21,  "#,##0"),
    ("Usable per basement floor", 25,  "#,##0.00"),
    ("Spaces before L-3/L-4",     21,  "#,##0.00"),
    ("Notes",                     46,  None),
]


def load_schedule(db_path):
    """designation -> (far, coverage fraction), plus the coefficient defaults."""
    if not os.path.exists(db_path):
        sys.exit(f"not found: {db_path}\nRun build_db.py first.")
    con = sqlite3.connect(db_path)
    sched = {}
    for des, far, cov in con.execute("SELECT designation, far, coverage FROM land_use"):
        sched[str(des).strip()] = (far, cov)
    coef = {k: v for k, v in con.execute("SELECT key, value FROM coefficient")}
    meta = {k: v for k, v in con.execute("SELECT key, value FROM meta")}
    con.close()
    return sched, coef, meta


def read_layer(gdb, layer):
    try:
        import pyogrio
        import pyogrio.raw as praw
    except ImportError:
        sys.exit("Reading a geodatabase needs GDAL via pyogrio:  pip install pyogrio")

    available = {name for name, _ in pyogrio.list_layers(gdb)}
    if layer not in available:
        sys.exit(f"layer {layer!r} not in {os.path.basename(gdb)}; "
                 f"layers present: {sorted(available)}")

    present = set(pyogrio.read_info(gdb, layer=layer)["fields"])
    if AREA_FIELD not in present:
        sys.exit(f"{layer} has no {AREA_FIELD} field, so there is no plot area to read from. "
                 f"Fields present: {sorted(present)}")
    cols = [c for c in WANTED if c in present]
    meta, _, _, data = praw.read(gdb, layer=layer, columns=cols, read_geometry=False)
    # praw.read returns the columns in its own order, so key by the names it gives back
    fields = {name: arr for name, arr in zip(meta["fields"], data)}
    count = len(next(iter(fields.values()))) if fields else 0
    missing = [c for c in WANTED if c not in present]
    return fields, count, missing


def cell(fields, name, i):
    arr = fields.get(name)
    if arr is None:
        return None
    v = arr[i]
    if v is None:
        return None
    if isinstance(v, float) and math.isnan(v):
        return None
    s = str(v).strip()
    if s in ("", "None", "nan", "<NA>"):
        return None
    return v


def as_float(v):
    """Numbers arrive as strings in this layer, and 'N/A' is a real value."""
    if v is None:
        return None
    try:
        f = float(str(v).strip())
    except (TypeError, ValueError):
        return None
    return None if math.isnan(f) else f


def derive(fields, i, sched, args):
    """One plot -> the derivation, plus notes on anything assumed."""
    notes = []
    area = as_float(cell(fields, AREA_FIELD, i))
    code = cell(fields, "DevCode", i)
    code = str(code).strip() if code is not None else None

    # FAR: the layer's own value first, the schedule second. A FAR of 0 is a real
    # value meaning no buildable area, distinct from 'N/A' meaning none published,
    # so only None falls through to the schedule.
    far = as_float(cell(fields, "DevCode_FAR", i))
    far_src = "geodatabase"
    if far is None:
        sched_far = sched.get(code, (None, None))[0] if code else None
        if sched_far is not None:
            far, far_src = float(sched_far), "Code schedule"
            notes.append("FAR taken from the Code; the geodatabase publishes none")
        else:
            far_src = "none"
            notes.append("no FAR published in either source, so no GFA could be derived")
    elif far == 0:
        notes.append("FAR is zero: no buildable area on this plot")

    # Coverage only exists in the schedule
    cov = sched.get(code, (None, None))[1] if code else None
    if cov is not None:
        cov_pct, cov_src = float(cov) * 100.0, "Code schedule"
    else:
        cov_pct, cov_src = float(args.coverage_default), "fallback default"
        notes.append(f"coverage not published for this designation, "
                     f"defaulted to {cov_pct:g}%")
    if code and code not in sched:
        notes.append(f"designation {code!r} is not in the Code schedule")

    have = area is not None
    max_gfa = area * far if (have and far is not None) else None
    cover = area * cov_pct / 100.0 if have else None
    gla = max_gfa * args.gla / 100.0 if max_gfa is not None else None
    acts = math.floor(gla / args.unit) if (gla is not None and args.unit > 0) else None

    # cross-check against the layer's own Max GFA
    gdb_gfa = as_float(cell(fields, "DevCode_MaxGFA", i))
    if max_gfa is not None and gdb_gfa is not None:
        if abs(max_gfa - gdb_gfa) <= max(0.05, abs(gdb_gfa) * 1e-6):
            check = "match"
        else:
            check = f"differs by {max_gfa - gdb_gfa:+.2f}"
            notes.append("computed Max GFA disagrees with the geodatabase field")
    elif max_gfa is not None:
        check = "no gdb value"
    else:
        check = ""

    # Parking capacity rather than a parking figure: how many spaces the open
    # ground outside the footprint holds, and how much a basement floor gives.
    # The number of spaces *required* needs an activity schedule, which no
    # geodatabase field carries.
    open_area = max(0.0, area - cover) if (have and cover is not None) else None
    open_spaces = math.floor(open_area / args.space) \
        if (open_area is not None and args.space > 0) else None
    per_floor = area * args.floor_use / 100.0 if have else None
    if code and code.upper() == "CR":
        notes.append("Community Retail: 100% coverage leaves no open ground, so all "
                     "parking goes to basement")

    # The cap is a ratio against GFA, so the ceiling in spaces is exact even
    # though the demand that would meet it is not knowable here.
    cap_spaces = max_gfa * PARK_CAP / 100.0 if max_gfa is not None else None
    if cap_spaces is not None:
        if area < SMALL_PLOT:
            notes.append(f"under {SMALL_PLOT:,.0f} m2: limitation L-4 bars any use "
                         f"demanding more than {cap_spaces:,.2f} spaces "
                         f"({PARK_CAP:g} per 100 m2 of GFA)")
        else:
            notes.append(f"L-3: demand over {cap_spaces:,.2f} spaces "
                         f"({PARK_CAP:g} per 100 m2 of GFA) is permitted but becomes "
                         f"the developer's responsibility, with no additional GFA")

    return [
        i + 1,
        cell(fields, "PLOTNUMBER", i),
        cell(fields, "SectorNumber_PlotID", i),
        cell(fields, "DISTRICTENG", i),
        cell(fields, "PRIMARYUSEENGDESC", i),
        code,
        cell(fields, "DevCode_Category", i),
        area, far, far_src,
        max_gfa, gdb_gfa, check,
        cov_pct / 100.0, cov_src, cover,
        gla, args.unit, acts,
        open_area, open_spaces, per_floor, cap_spaces,
        "; ".join(notes),
    ]


def write_xlsx(rows, out, args, sched_meta, gdb, layer, missing, counts):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("Writing .xlsx needs openpyxl:  pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Plots"

    head_font = Font(bold=True, size=9, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="0E6394")
    thin = Side(style="thin", color="BFCEDA")
    warn_fill = PatternFill("solid", fgColor="FDF3E2")

    for c, (name, width, _) in enumerate(COLUMNS, start=1):
        cl = ws.cell(row=1, column=c, value=name)
        cl.font = head_font
        cl.fill = head_fill
        cl.alignment = Alignment(vertical="center", wrap_text=True)
        cl.border = Border(bottom=thin)
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "C2"

    for r, row in enumerate(rows, start=2):
        no_far = row[8] is None
        for c, ((_, _, fmt), val) in enumerate(zip(COLUMNS, row), start=1):
            cl = ws.cell(row=r, column=c, value=val)
            if fmt:
                cl.number_format = fmt
            cl.border = Border(bottom=thin)
            if no_far:
                cl.fill = warn_fill
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"

    # what the numbers were produced with
    p = wb.create_sheet("Parameters")
    p.column_dimensions["A"].width = 34
    p.column_dimensions["B"].width = 62
    for k, v in [
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Geodatabase", os.path.basename(os.path.normpath(gdb))),
        ("Layer", layer),
        ("Plot area field", AREA_FIELD),
        ("Plots read", counts["total"]),
        ("Plots with a FAR", counts["with_far"]),
        ("Plots without a FAR", counts["no_far"]),
        ("Plots with a FAR of zero", counts["zero_far"]),
        ("", ""),
        ("GLA share of Max GFA", f"{args.gla:g}%"),
        ("Unit area per activity", f"{args.unit:g} m2"),
        ("Coverage fallback default", f"{args.coverage_default:g}%"),
        ("Area per parking space", f"{args.space:g} m2"),
        ("Usable per basement floor", f"{args.floor_use:g}%"),
        ("Parking demand cap (L-3/L-4)", f"{PARK_CAP:g} spaces per 100 m2 of GFA"),
        ("L-4 plot size threshold", f"{SMALL_PLOT:,.0f} m2"),
        ("", ""),
        ("Designation schedule", sched_meta.get("designation_file", "-")),
        ("Designations in schedule", sched_meta.get("designations", "-")),
        ("ITC rate source", sched_meta.get("source_file", "-")),
        ("Database built", sched_meta.get("built", "-")),
    ]:
        row = p.max_row + 1
        p.cell(row=row, column=1, value=k).font = Font(bold=bool(k), size=10)
        p.cell(row=row, column=2, value=v)
    if missing:
        row = p.max_row + 2
        p.cell(row=row, column=1, value="Fields absent from the layer").font = Font(bold=True, size=10)
        p.cell(row=row, column=2, value=", ".join(missing))

    n = wb.create_sheet("Notes")
    n.column_dimensions["A"].width = 108
    for line in [
        "How each column was produced",
        "",
        "Max GFA          = plot area x FAR",
        "Plot coverage    = plot area x max plot coverage%   (the GFA allowed on the ground floor)",
        "GLA              = Max GFA x GLA share",
        "Max activities   = floor(GLA / unit area)",
        "",
        "FAR comes from the layer's DevCode_FAR where it has one, otherwise from the Code",
        "schedule by DevCode. The FAR source column says which was used. Rows with no FAR from",
        "either source are shaded and carry no GFA.",
        "",
        "Max plot coverage is not a geodatabase field, so it always comes from the Code schedule.",
        "Where the Code publishes none it falls back to the default on the Parameters sheet, and",
        "the row says so.",
        "",
        "Parking is NOT calculated per plot. The number of spaces required depends on which",
        "activities occupy the plot, and no geodatabase field carries that. The parking columns",
        "give capacity instead:",
        "",
        "  Open ground           = plot area - plot coverage",
        "  Spaces on open ground = floor(open ground / area per space)",
        "  Usable per basement floor = plot area x usable share  (a basement is not held to coverage)",
        "  Spaces before L-3/L-4 = Max GFA x 2 / 100",
        "",
        "Spaces beyond what the open ground holds go to basement, at the area per space, and a",
        "basement floor yields only its usable share -- so floors = ceil(basement area /",
        "usable per floor). Community Retail is the case that forces this: the Code gives it 100%",
        "plot coverage, so there is no open ground at all.",
        "",
        "Limitations L-3 and L-4 cap parking demand at 2 spaces per 100 sqm of GFA. The ceiling",
        "that puts on a plot is exact, so it is given as a column; whether a scheme reaches it is",
        "not, because that needs the activity schedule. Below 1,600 sqm L-4 bars a use that",
        "exceeds the ceiling; at or above, L-3 permits it but makes the excess the developer's",
        "responsibility and grants no additional GFA.",
        "",
        "Unconfirmed with DMT at the time of writing:",
        "  - the GLA share (75% by default)",
        "Treat the derived figures as working numbers, not cleared outputs.",
    ]:
        r = n.max_row + 1
        c = n.cell(row=r, column=1, value=line)
        if line and not line.startswith(" ") and line.endswith("produced"):
            c.font = Font(bold=True, size=11)
        elif line.startswith("Unconfirmed"):
            c.font = Font(bold=True, size=10)

    wb.save(out)


def main():
    ap = argparse.ArgumentParser(
        description="Run the DCR derivation over a geodatabase plot layer, to Excel.")
    ap.add_argument("gdb", help="path to the .gdb directory")
    ap.add_argument("-o", "--out", help="output .xlsx (default: alongside the .gdb)")
    ap.add_argument("--layer", default="UDM_Plot", help="plot layer name (default UDM_Plot)")
    ap.add_argument("--gla", type=float, default=None, help="GLA share of Max GFA, %%")
    ap.add_argument("--unit", type=float, default=60.0, help="unit area per activity, m2")
    ap.add_argument("--coverage-default", type=float, default=None,
                    help="coverage %% where the Code publishes none")
    ap.add_argument("--space", type=float, default=32.5, help="area per parking space, m2")
    ap.add_argument("--floor-use", type=float, default=75.0,
                    help="usable share of a basement floor, %%")
    args = ap.parse_args()

    gdb = os.path.normpath(args.gdb)
    if not os.path.isdir(gdb):
        sys.exit(f"not a geodatabase directory: {gdb}")

    sched, coef, sched_meta = load_schedule(DB)
    if args.gla is None:
        args.gla = float(coef.get("gla_pct", 75))
    if args.coverage_default is None:
        args.coverage_default = float(coef.get("coverage_pct_default", 60))

    fields, count, missing = read_layer(gdb, args.layer)
    rows = [derive(fields, i, sched, args) for i in range(count)]

    counts = {
        "total": count,
        "with_far": sum(1 for r in rows if r[8] is not None),
        "no_far": sum(1 for r in rows if r[8] is None),
        "zero_far": sum(1 for r in rows if r[8] == 0),
    }
    out = args.out or os.path.join(os.path.dirname(gdb) or ".",
                                   os.path.basename(gdb).replace(".gdb", "") + "_DCR.xlsx")
    write_xlsx(rows, out, args, sched_meta, gdb, args.layer, missing, counts)

    print(f"{out}")
    print(f"  {count} plots from {args.layer}  ({counts['with_far']} with a FAR, "
          f"of which {counts['zero_far']} are zero; {counts['no_far']} with none published)")
    mism = [r for r in rows if isinstance(r[12], str) and r[12].startswith("differs")]
    print(f"  Max GFA cross-check against the layer: "
          f"{sum(1 for r in rows if r[12] == 'match')} match, {len(mism)} differ")
    unknown = sorted({r[5] for r in rows if r[5] and r[5] not in sched})
    if unknown:
        print(f"  designations not in the Code schedule: {', '.join(unknown)}")


if __name__ == "__main__":
    main()
